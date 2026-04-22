import os
import re
import sys
import math
import time
import shutil
import unicodedata
import webbrowser
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlencode, quote

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

try:
    from PIL import Image, ImageTk
except ImportError:
    Image = None
    ImageTk = None

try:
    import pytesseract
except ImportError:
    pytesseract = None


APP_TITLE = "Artprice Link Generator"
DEFAULT_DATE_FROM = "2016-01-01"
DEFAULT_CATEGORY_ID = "2"
DEFAULT_SORT = "datesale_desc"

DEFAULT_ARTIST_ID_PATH = "/Users/theodoredonson/secondstatesite/artpricelinkgen/ARTIST IDs.xlsx"
DEFAULT_DAUMIER_IMAGE_PATH = "/Users/theodoredonson/PycharmProjects/APRTPRICE LINK GEN2/daumier_smoking_guy.png"

FALLBACK_ARTIST_ID_FILENAMES = [
    "ARTIST IDs.xlsx",
    "ARTIST IDs - ARTIST IDs.csv",
    "ARTIST IDs Harvested v2.xlsx",
]

UNKNOWN_ARTIST_MESSAGE = "Sorry :( The artist ID # is not known!"

ARTPRICE_LINK_COLUMN = "Artprice Link"
ARTIST_COLUMN = "Extracted Artist"
TITLE_COLUMN = "Extracted Title"
ARTIST_ID_COLUMN = "Artprice Artist ID"
KEYWORD_COLUMN = "Artprice Keyword"
STATUS_COLUMN = "Artprice Status"
SOURCE_MODE_COLUMN = "Artprice Source Mode"
MISSING_SEARCH_COLUMN = "Artprice Artist Search"

BG = "#f6f2ea"
PANEL = "#fbf8f2"
TEXT = "#2a2118"
SUBTLE = "#6c5a45"
GOLD_1 = "#8c6a1b"
GOLD_2 = "#b88928"
GOLD_3 = "#d7b55a"
GOLD_4 = "#f6df9a"
GOLD_SHADOW = "#9b7a22"
PINK_1 = "#f5d7e1"
PINK_2 = "#dba8bb"
LINK_BLUE = "#1d4ed8"


@dataclass
class ExtractedListing:
    artist: str
    title: str
    source_url: str = ""
    raw_heading: str = ""


def blend_hex(a, b, t):
    def to_rgb(hex_color):
        hex_color = hex_color.lstrip("#")
        return tuple(int(hex_color[i:i + 2], 16) for i in (0, 2, 4))
    ar, ag, ab = to_rgb(a)
    br, bg, bb = to_rgb(b)
    rr = int(ar + (br - ar) * t)
    rg = int(ag + (bg - ag) * t)
    rb = int(ab + (bb - ab) * t)
    return f"#{rr:02x}{rg:02x}{rb:02x}"


class ArtistIdLookup:
    def __init__(self):
        self.df = None
        self.path = None
        self.normalized_map = {}

    @staticmethod
    def strip_parenthetical(value: str) -> str:
        return re.sub(r"\([^)]*\)", "", str(value or "")).strip()

    @staticmethod
    def reorder_name(value: str) -> str:
        value = str(value or "").strip()
        if "," in value and value.count(",") == 1:
            last, first = [part.strip() for part in value.split(",", 1)]
            if first and last:
                value = f"{first} {last}"
        return value

    @classmethod
    def clean_display_name(cls, value: str) -> str:
        value = cls.reorder_name(cls.strip_parenthetical(value))
        value = re.sub(
            r"\b(?:after|attributed to|attr\.?|school of|manner of|circle of|follower of)\b",
            "",
            value,
            flags=re.I,
        )
        value = re.sub(r"[^A-Za-z0-9' .,\-]+", " ", value)
        value = re.sub(r"\s+", " ", value).strip(" ,;:-")
        return value

    @staticmethod
    def normalize_name(value: str) -> str:
        value = unicodedata.normalize("NFKD", str(value or ""))
        value = "".join(ch for ch in value if not unicodedata.combining(ch))
        value = value.lower().replace("&", " and ")
        value = re.sub(r"[^a-z0-9]+", " ", value)
        return re.sub(r"\s+", " ", value).strip()

    def load_file(self, path: str):
        if not path:
            raise ValueError("No artist ID file was provided.")
        if path.lower().endswith(".csv"):
            df = pd.read_csv(path)
        else:
            df = pd.read_excel(path)

        df.columns = [str(c).strip() for c in df.columns]
        if not {"Artist", "ID"}.issubset(df.columns):
            raise ValueError("Artist ID file must contain 'Artist' and 'ID' columns.")

        df = df[["Artist", "ID"]].dropna(subset=["Artist", "ID"]).copy()
        df["Artist"] = df["Artist"].map(self.clean_display_name)
        df["ID"] = df["ID"].astype(str).str.extract(r"(\d+)", expand=False).fillna("")
        df = df[(df["Artist"] != "") & (df["ID"] != "")].drop_duplicates(subset=["Artist"], keep="first")

        self.df = df
        self.path = path
        self.normalized_map = {}
        for _, row in df.iterrows():
            norm = self.normalize_name(row["Artist"])
            if norm and norm not in self.normalized_map:
                self.normalized_map[norm] = row["ID"]

    def lookup(self, artist_name: str):
        cleaned = self.clean_display_name(artist_name)
        norm = self.normalize_name(cleaned)
        if not norm:
            return None
        if norm in self.normalized_map:
            return self.normalized_map[norm]
        norm_words = set(norm.split())
        for known_norm, artist_id in self.normalized_map.items():
            known_words = set(known_norm.split())
            if norm_words and norm_words == known_words:
                return artist_id
        for known_norm, artist_id in self.normalized_map.items():
            if norm in known_norm or known_norm in norm:
                return artist_id
        return None


class ImageListingExtractor:
    def __init__(self):
        pass

    @staticmethod
    def _clean_text(value: str) -> str:
        value = str(value or "")
        value = unicodedata.normalize("NFKD", value)
        value = "".join(ch for ch in value if not unicodedata.combining(ch))
        value = value.replace("—", "-").replace("–", "-")
        value = re.sub(r"\s+", " ", value)
        return value.strip()

    def clean_artist_name(self, value: str) -> str:
        value = self._clean_text(value)
        value = re.sub(r"\([^)]*\)", "", value)
        value = re.sub(r"\b(?:after|attributed to|attr\.?|school of|manner of|circle of|follower of)\b", "", value, flags=re.I)
        value = re.sub(r"\b(?:american|british|french|german|spanish|italian|japanese|chinese|mexican|canadian)\b", "", value, flags=re.I)
        value = re.sub(r"\b(?:born|b\.)\s*\d{4}\b", "", value, flags=re.I)
        value = re.sub(r"\b\d{4}\s*-\s*\d{4}\b", "", value)
        value = re.sub(r"[^A-Za-z0-9' .,\-]+", " ", value)
        value = re.sub(r"\s*;.*$", "", value)
        value = re.sub(r"\s+/.*$", "", value)
        value = re.sub(r"\s+", " ", value).strip(" ,;:-")
        if "," in value and value.count(",") == 1:
            last, first = [p.strip() for p in value.split(",", 1)]
            if first and last:
                value = f"{first} {last}"
        return value

    def clean_title(self, value: str) -> str:
        value = self._clean_text(value)
        value = re.sub(r"\([^)]*\)", "", value)
        value = re.sub(r"\[[^\]]*\]", "", value)
        value = re.sub(r",?\s*(?:from|from the|from an?)\s+[^,;]+(?:series|suite|set|portfolio|album)?", "", value, flags=re.I)
        value = re.sub(r",?\s*(?:series|suite|set|portfolio|album)\s+[^,;]+$", "", value, flags=re.I)
        value = re.sub(r",?\s*(19|20)\d{2}$", "", value)
        value = re.sub(r"\b(?:signed|dated|numbered|framed|sheet|image size|estimate|dimensions)\b.*$", "", value, flags=re.I)
        value = re.sub(r"\s+", " ", value).strip(" ,;:-")
        return value

    def _ocr_text(self, image_path: str) -> str:
        if Image is None or pytesseract is None:
            raise RuntimeError(
                "OCR dependencies are missing. Install Pillow and pytesseract, and install the Tesseract app on your Mac."
            )
        if shutil.which("tesseract") is None:
            raise RuntimeError(
                "Tesseract OCR is not installed on this Mac. Install it with: brew install tesseract"
            )
        img = Image.open(image_path)
        text = pytesseract.image_to_string(img)
        return text

    def extract_from_image(self, image_path: str) -> ExtractedListing:
        raw = self._ocr_text(image_path)
        text = self._clean_text(raw)

        lines = [self._clean_text(line) for line in raw.splitlines()]
        lines = [line for line in lines if line and len(line) > 2]

        artist = ""
        title = ""

        artist_patterns = [
            r"([A-Z][A-Za-z' .\-]+)\s*\((?:[^)]*)\)",
            r"([A-Z][A-Za-z' .\-]+)\s+\d{4}\s*-\s*\d{4}",
            r"([A-Z][A-Za-z' .\-]+)\s+(?:American|British|French|German|Spanish|Italian|Japanese|Chinese|Mexican|Canadian)",
        ]

        for line in lines[:12]:
            for pattern in artist_patterns:
                m = re.search(pattern, line, flags=re.I)
                if m:
                    artist = self.clean_artist_name(m.group(1))
                    break
            if artist:
                break

        if not artist:
            for line in lines[:12]:
                clean = self.clean_artist_name(line)
                words = clean.split()
                if 2 <= len(words) <= 4 and all(w[:1].isalpha() for w in words if w):
                    artist = clean
                    break

        title_candidates = []
        for line in lines[:20]:
            clean = self.clean_title(line)
            if not clean:
                continue
            if artist and self.clean_artist_name(clean).lower() == artist.lower():
                continue
            if re.search(r"(estimate|dimensions|sheet|image size|signed|framed|edition)", clean, flags=re.I):
                continue
            if len(clean.split()) >= 1:
                title_candidates.append(clean)

        if title_candidates:
            title = title_candidates[0]

        if not artist or not title:
            raise ValueError("Could not confidently read artist/title from the uploaded image.")

        return ExtractedListing(artist=artist, title=title, source_url=image_path, raw_heading=text[:500])


class ArtpriceURLBuilder:
    @staticmethod
    def resolve_mode(exact_match: bool, all_terms: bool) -> str:
        if exact_match:
            return "exact"
        if all_terms:
            return "all_terms"
        return "plain"

    @staticmethod
    def clean_title_for_keyword(title: str) -> str:
        value = str(title or "").strip()
        value = unicodedata.normalize("NFKD", value)
        value = "".join(ch for ch in value if not unicodedata.combining(ch))
        value = re.sub(r"\([^)]*\)", "", value)
        value = re.sub(r"\[[^\]]*\]", "", value)
        value = re.sub(r",?\s*(?:from|from the|from an?)\s+[^,;]+(?:series|suite|set|portfolio|album)?", "", value, flags=re.I)
        value = re.sub(r",?\s*(?:series|suite|set|portfolio|album)\s+[^,;]+$", "", value, flags=re.I)
        value = re.sub(r",?\s*(19|20)\d{2}$", "", value)
        value = re.sub(r"\s+", " ", value).strip(" ,;:-")
        return value

    @classmethod
    def build_keyword(cls, title: str, exact_match: bool, all_terms: bool) -> str:
        title = cls.clean_title_for_keyword(title)
        parts = re.findall(r"[A-Za-z0-9']+", title)
        phrase = " ".join(parts).strip()
        if not phrase:
            return title
        if exact_match:
            return phrase
        if all_terms:
            return "-".join(parts)
        return phrase

    @classmethod
    def build_url(cls, artist_id: str, title: str, exact_match: bool, all_terms: bool) -> str:
        params = {
            "dt_from": DEFAULT_DATE_FROM,
            "exact_match": "1" if exact_match else "0",
            "idartist": str(artist_id),
            "idcategory": DEFAULT_CATEGORY_ID,
            "keyword": cls.build_keyword(title, exact_match, all_terms),
            "p": "1",
            "sort": DEFAULT_SORT,
        }
        return "https://www.artprice.com/lots/search?" + urlencode(params)

    @classmethod
    def build_url_without_artist(cls, title: str, exact_match: bool, all_terms: bool) -> str:
        params = {
            "dt_from": DEFAULT_DATE_FROM,
            "exact_match": "1" if exact_match else "0",
            "idcategory": DEFAULT_CATEGORY_ID,
            "keyword": cls.build_keyword(title, exact_match, all_terms),
            "p": "1",
            "sort": DEFAULT_SORT,
        }
        return "https://www.artprice.com/lots/search?" + urlencode(params)

    @staticmethod
    def build_artist_search_url(artist_name: str) -> str:
        return f"https://www.artprice.com/artists/search?keyword={quote(artist_name)}"


class BatchProcessor:
    def __init__(self, extractor: ImageListingExtractor, lookup: ArtistIdLookup, logger):
        self.extractor = extractor
        self.lookup = lookup
        self.logger = logger

    @staticmethod
    def detect_column(df: pd.DataFrame, names):
        lower_map = {str(col).strip().lower(): col for col in df.columns}
        for name in names:
            if name in lower_map:
                return lower_map[name]
        return None

    def detect_artist_column(self, df: pd.DataFrame):
        return self.detect_column(df, ["artist", "artists", "maker", "artist name", "name"])

    def detect_title_column(self, df: pd.DataFrame):
        return self.detect_column(df, ["title", "artwork title", "work title", "lot title", "object title"])

    def process_workbook(self, input_path: str, exact_match: bool, all_terms: bool, progress_callback=None):
        df = pd.read_excel(input_path)
        if df.empty:
            raise ValueError("The spreadsheet is empty.")

        artist_col = self.detect_artist_column(df)
        title_col = self.detect_title_column(df)
        if not (artist_col and title_col):
            raise ValueError("Could not find Artist and Title columns in that spreadsheet.")

        result_df = df.copy()
        result_df = result_df.astype(object)

        for col in [ARTIST_COLUMN, TITLE_COLUMN, ARTIST_ID_COLUMN, KEYWORD_COLUMN, ARTPRICE_LINK_COLUMN, STATUS_COLUMN, SOURCE_MODE_COLUMN]:
            if col not in result_df.columns:
                result_df[col] = ""

        missing_artists = set()
        total = len(result_df)

        for row_num, idx in enumerate(result_df.index, start=1):
            artist = str(result_df.at[idx, artist_col]).strip() if pd.notna(result_df.at[idx, artist_col]) else ""
            title = str(result_df.at[idx, title_col]).strip() if pd.notna(result_df.at[idx, title_col]) else ""

            if not artist or not title:
                result_df.at[idx, STATUS_COLUMN] = "Skipped: missing artist/title"
                result_df.at[idx, SOURCE_MODE_COLUMN] = "sheet"
                if progress_callback:
                    progress_callback(row_num, total, result_df.at[idx, STATUS_COLUMN])
                continue

            listing = ExtractedListing(
                artist=self.extractor.clean_artist_name(artist),
                title=self.extractor.clean_title(title),
                source_url=input_path,
            )

            result_df.at[idx, ARTIST_COLUMN] = listing.artist
            result_df.at[idx, TITLE_COLUMN] = listing.title
            result_df.at[idx, SOURCE_MODE_COLUMN] = "sheet"

            artist_id = self.lookup.lookup(listing.artist)
            keyword = ArtpriceURLBuilder.build_keyword(listing.title, exact_match, all_terms)

            if not artist_id:
                link = ArtpriceURLBuilder.build_url_without_artist(listing.title, exact_match, all_terms)
                result_df.at[idx, ARTIST_ID_COLUMN] = ""
                result_df.at[idx, KEYWORD_COLUMN] = keyword
                result_df.at[idx, ARTPRICE_LINK_COLUMN] = link
                result_df.at[idx, STATUS_COLUMN] = UNKNOWN_ARTIST_MESSAGE
                missing_artists.add(listing.artist)
                self.logger(f"Missing artist ID match: {listing.artist}")
                self.logger(f"Generated fallback batch link without artist ID for: {listing.artist}")
                if progress_callback:
                    progress_callback(row_num, total, result_df.at[idx, STATUS_COLUMN])
                continue

            link = ArtpriceURLBuilder.build_url(artist_id, listing.title, exact_match, all_terms)
            result_df.at[idx, ARTIST_ID_COLUMN] = str(artist_id)
            result_df.at[idx, KEYWORD_COLUMN] = keyword
            result_df.at[idx, ARTPRICE_LINK_COLUMN] = link
            result_df.at[idx, STATUS_COLUMN] = "OK"
            if progress_callback:
                progress_callback(row_num, total, "OK")

        output_path = self.build_output_path(input_path)
        result_df.to_excel(output_path, index=False)
        format_output_workbook(output_path, link_header=ARTPRICE_LINK_COLUMN)
        missing_file = self.export_missing_artists_file(missing_artists)
        return output_path, missing_file, artist_col, title_col, result_df

    @staticmethod
    def build_output_path(input_path: str) -> str:
        p = Path(input_path)
        return str(p.with_name(f"{p.stem}_with_Artprice_Links{p.suffix}"))

    def export_missing_artists_file(self, missing_artists):
        if not missing_artists:
            return None
        desktop = Path.home() / "Desktop"
        try:
            desktop.mkdir(parents=True, exist_ok=True)
        except Exception:
            desktop = Path.cwd()

        output = desktop / "Missing Artist Artprice IDs.xlsx"
        rows = []
        for artist in sorted({self.extractor.clean_artist_name(a) for a in missing_artists if str(a).strip()}):
            rows.append(
                {
                    "Artist": artist,
                    "ID": "",
                    MISSING_SEARCH_COLUMN: ArtpriceURLBuilder.build_artist_search_url(artist),
                }
            )

        pd.DataFrame(rows).to_excel(output, index=False)
        format_output_workbook(str(output), link_header=MISSING_SEARCH_COLUMN)
        self.logger(f"Missing Artist Artprice IDs exported to desktop: {output}")
        return str(output)


class HyperlinkText(tk.Text):
    def __init__(self, master, **kwargs):
        super().__init__(master, **kwargs)
        self.config(cursor="arrow")
        self.tag_configure("link", foreground=LINK_BLUE, underline=True)
        self.links = {}

    def insert_link(self, label: str, url: str):
        start = self.index("end-1c")
        self.insert("end", label)
        end = self.index("end-1c")
        tag = f"link_{len(self.links)}"
        self.links[tag] = url
        self.tag_add("link", start, end)
        self.tag_add(tag, start, end)
        self.tag_bind(tag, "<Enter>", lambda e: self.config(cursor="hand2"))
        self.tag_bind(tag, "<Leave>", lambda e: self.config(cursor="arrow"))
        self.tag_bind(tag, "<Button-1>", lambda e, t=tag: webbrowser.open(self.links[t]))


class GoldButton(tk.Frame):
    def __init__(self, parent, text, command, width=None):
        super().__init__(parent, bg=GOLD_SHADOW, highlightthickness=0, bd=0)
        self.button = tk.Button(
            self,
            text=text,
            command=command,
            bg=GOLD_3,
            fg=TEXT,
            activebackground=GOLD_4,
            activeforeground=TEXT,
            relief="flat",
            bd=0,
            padx=14,
            pady=10,
            font=("Georgia", 11, "bold"),
            cursor="hand2",
            width=width,
        )
        self.button.pack(padx=(0, 2), pady=(0, 2))


class App:
    def __init__(self, master):
        self.master = master
        self.master.title(APP_TITLE)
        self.master.geometry("1240x920")
        self.master.minsize(1120, 840)
        self.master.configure(bg=BG)

        self.lookup = ArtistIdLookup()
        self.extractor = ImageListingExtractor()
        self.last_generated_link = ""
        self.last_exported_file = None
        self.last_missing_file = None
        self.batch_input_path = ""
        self.image_input_path = ""
        self.batch_processor = BatchProcessor(self.extractor, self.lookup, self.log)
        self.daumier_img_ref = None

        self._build_fonts()
        self._build_ui()
        self._autoload_artist_ids()
        self._load_daumier_image()
        self._animate_banner(0)

    def log(self, message: str):
        stamp = time.strftime("%H:%M:%S")
        text = f"[{stamp}] {message}"
        print(text)
        sys.stdout.flush()
        self.log_text.config(state="normal")
        self.log_text.insert("end", text + "\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _build_fonts(self):
        self.title_font = ("Baskerville", 30, "bold")
        self.section_font = ("Georgia", 15, "bold")
        self.label_font = ("Georgia", 12, "bold")
        self.body_font = ("Helvetica", 12)
        self.small_font = ("Helvetica", 10)

    def _build_ui(self):
        outer = tk.Frame(self.master, bg=BG)
        outer.pack(fill="both", expand=True, padx=18, pady=18)

        self.banner = tk.Canvas(outer, height=24, bg=BG, highlightthickness=0)
        self.banner.pack(fill="x", pady=(0, 10))

        header = tk.Frame(outer, bg=BG)
        header.pack(fill="x")

        header_left = tk.Frame(header, bg=BG)
        header_left.pack(side="left", fill="x", expand=True)

        tk.Label(header_left, text="Artprice Link Generator", bg=BG, fg=TEXT, font=self.title_font).pack(anchor="w")
        tk.Label(
            header_left,
            text="Single listing screenshot and batch spreadsheet Artprice link generation.",
            bg=BG,
            fg=SUBTLE,
            font=self.body_font,
        ).pack(anchor="w", pady=(4, 10))

        self.header_image_label = tk.Label(header, bg=BG)
        self.header_image_label.pack(side="right", padx=(12, 0))

        options_shell = tk.Frame(outer, bg=GOLD_SHADOW)
        options_shell.pack(fill="x", pady=(0, 14))
        options = tk.Frame(options_shell, bg=PANEL)
        options.pack(fill="both", expand=True, padx=(0, 3), pady=(0, 3))

        tk.Label(options, text="Search Options", bg=PANEL, fg=TEXT, font=self.section_font).pack(anchor="w", padx=12, pady=(10, 0))
        opt_row = tk.Frame(options, bg=PANEL)
        opt_row.pack(fill="x", pady=(10, 0), padx=12)

        self.exact_match_var = tk.BooleanVar(value=False)
        self.all_terms_var = tk.BooleanVar(value=False)

        tk.Checkbutton(
            opt_row,
            text="Exact match",
            variable=self.exact_match_var,
            bg=PANEL,
            fg=TEXT,
            activebackground=PANEL,
            selectcolor="#fff6dc",
            font=self.body_font,
        ).pack(side="left")

        tk.Checkbutton(
            opt_row,
            text="Search all terms individually",
            variable=self.all_terms_var,
            bg=PANEL,
            fg=TEXT,
            activebackground=PANEL,
            selectcolor="#fff6dc",
            font=self.body_font,
        ).pack(side="left", padx=(18, 0))

        tk.Label(
            options,
            text="Exact match checked = plain title with exact_match 1. Search all terms individually = dashed terms with exact_match 0. Neither checked = plain title with exact_match 0.",
            bg=PANEL,
            fg=SUBTLE,
            font=self.small_font,
            wraplength=1020,
            justify="left",
        ).pack(anchor="w", padx=12, pady=(8, 12))

        notebook_shell = tk.Frame(outer, bg=GOLD_SHADOW)
        notebook_shell.pack(fill="both", expand=True)
        notebook_frame = tk.Frame(notebook_shell, bg=PANEL)
        notebook_frame.pack(fill="both", expand=True, padx=(0, 3), pady=(0, 3))

        style = ttk.Style(self.master)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TNotebook", background=PANEL, borderwidth=0)
        style.configure("TNotebook.Tab", font=("Georgia", 12, "bold"), padding=(18, 10), background="#efe7d3", foreground=TEXT)
        style.map("TNotebook.Tab", background=[("selected", PANEL)], foreground=[("selected", TEXT)])

        self.notebook = ttk.Notebook(notebook_frame)
        self.notebook.pack(fill="both", expand=True, padx=8, pady=8)

        self.image_tab = tk.Frame(self.notebook, bg=PANEL)
        self.batch_tab = tk.Frame(self.notebook, bg=PANEL)

        self.notebook.add(self.image_tab, text="Upload Photo of Auction Listing")
        self.notebook.add(self.batch_tab, text="Batch Spreadsheet")

        self._build_image_tab(self.image_tab)
        self._build_batch_tab(self.batch_tab)

        bottom = tk.Frame(outer, bg=BG)
        bottom.pack(fill="both", expand=False, pady=(14, 0))

        log_shell = tk.Frame(bottom, bg=GOLD_SHADOW)
        log_shell.pack(side="left", fill="both", expand=True, padx=(0, 8))
        log_panel = tk.Frame(log_shell, bg=PANEL)
        log_panel.pack(fill="both", expand=True, padx=(0, 3), pady=(0, 3))

        tk.Label(log_panel, text="Program Log", bg=PANEL, fg=TEXT, font=self.section_font).pack(anchor="w", padx=12, pady=(10, 0))
        self.log_text = tk.Text(log_panel, height=10, bg="#fffdf9", fg=TEXT, font=("Menlo", 10), relief="flat", bd=0, wrap="word")
        self.log_text.pack(fill="both", expand=True, padx=12, pady=(8, 12))
        self.log_text.config(state="disabled")

        status_shell = tk.Frame(bottom, bg=GOLD_SHADOW)
        status_shell.pack(side="left", fill="both", expand=True, padx=(8, 0))
        status_panel = tk.Frame(status_shell, bg=PANEL)
        status_panel.pack(fill="both", expand=True, padx=(0, 3), pady=(0, 3))

        tk.Label(status_panel, text="Status", bg=PANEL, fg=TEXT, font=self.section_font).pack(anchor="w", padx=12, pady=(10, 0))
        self.status_text = HyperlinkText(status_panel, height=10, bg="#fffdf9", fg=TEXT, font=self.body_font, relief="flat", bd=0, wrap="word")
        self.status_text.pack(fill="both", expand=True, padx=12, pady=(8, 12))
        self.status_text.insert("end", "Ready.\n")

    def _build_image_tab(self, parent):
        wrap = tk.Frame(parent, bg=PANEL)
        wrap.pack(fill="both", expand=True, padx=16, pady=16)

        tk.Label(wrap, text="Upload Photo of Auction Listing", bg=PANEL, fg=TEXT, font=self.section_font).pack(anchor="w")
        tk.Label(
            wrap,
            text="Upload a screenshot/photo of the auction listing. OCR will try to read artist and title from the image.",
            bg=PANEL,
            fg=SUBTLE,
            font=self.body_font,
            wraplength=1000,
            justify="left",
        ).pack(anchor="w", pady=(6, 10))

        file_row = tk.Frame(wrap, bg=PANEL)
        file_row.pack(fill="x")
        self.image_file_var = tk.StringVar()
        tk.Entry(file_row, textvariable=self.image_file_var, font=self.body_font, relief="flat", bd=0, bg="#fffdf9", fg=TEXT).pack(side="left", fill="x", expand=True, ipady=10)
        GoldButton(file_row, "Choose Image", self.choose_image_file).pack(side="left", padx=(10, 0))

        btn_row = tk.Frame(wrap, bg=PANEL)
        btn_row.pack(fill="x", pady=(12, 0))
        GoldButton(btn_row, "Generate Link", self.generate_link_from_image).pack(side="left")
        GoldButton(btn_row, "🌐 ↗ Open Link in Browser", self.open_link).pack(side="left", padx=(10, 0))
        GoldButton(btn_row, "Copy ArtPrice URL", self.copy_link).pack(side="left", padx=(10, 0))
        GoldButton(btn_row, "Clear", self.clear_image_tab).pack(side="left", padx=(10, 0))

        data_shell = tk.Frame(wrap, bg=GOLD_SHADOW)
        data_shell.pack(fill="both", expand=True, pady=(14, 0))
        data_panel = tk.Frame(data_shell, bg=PANEL)
        data_panel.pack(fill="both", expand=True, padx=(0, 3), pady=(0, 3))

        left = tk.Frame(data_panel, bg=PANEL)
        left.pack(side="left", fill="both", expand=True, padx=(12, 8), pady=12)
        right = tk.Frame(data_panel, bg=PANEL)
        right.pack(side="left", fill="both", expand=True, padx=(8, 12), pady=12)

        tk.Label(left, text="Extracted Data", bg=PANEL, fg=TEXT, font=self.section_font).pack(anchor="w")
        self.artist_var = tk.StringVar()
        self.title_var = tk.StringVar()
        self.artist_id_var = tk.StringVar()
        self.keyword_var = tk.StringVar()

        for label, var in [
            ("Artist", self.artist_var),
            ("Title", self.title_var),
            ("Artist ID", self.artist_id_var),
            ("Keyword", self.keyword_var),
        ]:
            row = tk.Frame(left, bg=PANEL)
            row.pack(fill="x", pady=6)
            tk.Label(row, text=label, width=12, anchor="w", bg=PANEL, fg=TEXT, font=self.label_font).pack(side="left")
            tk.Entry(row, textvariable=var, font=self.body_font, relief="flat", bd=0, bg="#fffdf9", fg=TEXT).pack(side="left", fill="x", expand=True, ipady=7)

        tk.Label(right, text="Result", bg=PANEL, fg=TEXT, font=self.section_font).pack(anchor="w")
        self.output = HyperlinkText(right, height=16, bg="#fffdf9", fg=TEXT, font=self.body_font, relief="flat", bd=0, wrap="word")
        self.output.pack(fill="both", expand=True, pady=(10, 0))

    def _build_batch_tab(self, parent):
        wrap = tk.Frame(parent, bg=PANEL)
        wrap.pack(fill="both", expand=True, padx=16, pady=16)

        tk.Label(
            wrap,
            text="Choose an auction spreadsheet. Artist and Title are preferred.",
            bg=PANEL,
            fg=SUBTLE,
            font=self.body_font,
            wraplength=1040,
            justify="left",
        ).pack(anchor="w")

        row = tk.Frame(wrap, bg=PANEL)
        row.pack(fill="x", pady=(10, 10))
        self.batch_file_var = tk.StringVar()
        tk.Entry(row, textvariable=self.batch_file_var, font=self.body_font, relief="flat", bd=0, bg="#fffdf9", fg=TEXT).pack(side="left", fill="x", expand=True, ipady=10)
        GoldButton(row, "Select Auction Spreadsheet", self.choose_batch_file).pack(side="left", padx=(10, 0))

        self.generate_batch_button = GoldButton(wrap, "Generate updated spreadsheet with artprice links", self.process_batch_file)
        self.generate_batch_button.pack(anchor="w", pady=(0, 10))
        self.generate_batch_button.pack_forget()

        log_shell = tk.Frame(wrap, bg=GOLD_SHADOW)
        log_shell.pack(fill="both", expand=True)
        log_panel = tk.Frame(log_shell, bg=PANEL)
        log_panel.pack(fill="both", expand=True, padx=(0, 3), pady=(0, 3))
        tk.Label(log_panel, text="Batch Output", bg=PANEL, fg=TEXT, font=self.section_font).pack(anchor="w", padx=12, pady=(10, 0))
        self.batch_log = tk.Text(log_panel, height=18, bg="#fffdf9", fg=TEXT, font=("Menlo", 10), relief="flat", bd=0, wrap="word")
        self.batch_log.pack(fill="both", expand=True, padx=12, pady=(8, 12))

    def _autoload_artist_ids(self):
        candidates = [DEFAULT_ARTIST_ID_PATH]
        base = os.path.dirname(os.path.abspath(__file__))
        candidates.extend(os.path.join(base, name) for name in FALLBACK_ARTIST_ID_FILENAMES)

        for path in candidates:
            if os.path.exists(path):
                self.lookup.load_file(path)
                self.log(f"Loaded {os.path.basename(path)} for link gen program")
                self._set_status_message(f"Loaded default artist ID file: {path}")
                return

        self.log("Default artist ID file was not found.")
        self._set_status_message("Default artist ID file was not found. Place ARTIST IDs.xlsx at the project path.")

    def _load_daumier_image(self):
        if Image is None or ImageTk is None:
            return

        img_path = None
        candidates = [DEFAULT_DAUMIER_IMAGE_PATH]
        base = os.path.dirname(os.path.abspath(__file__))
        for name in [
            "daumier_smoking_guy.png",
            "daumier_smoking_guy.jpg",
            "daumier.png",
            "daumier.jpg",
        ]:
            candidates.append(os.path.join(base, name))

        for path in candidates:
            if os.path.exists(path):
                img_path = path
                break

        if not img_path:
            return

        try:
            img = Image.open(img_path)
            img.thumbnail((240, 360))
            self.daumier_img_ref = ImageTk.PhotoImage(img)
            self.header_image_label.config(image=self.daumier_img_ref)
        except Exception as exc:
            self.log(f"Could not load Daumier image: {exc}")

    def _set_status_message(self, text, link_label=None, link_url=None):
        self.status_text.delete("1.0", "end")
        self.status_text.insert("end", text)
        if link_label and link_url:
            self.status_text.insert("end", "\n")
            self.status_text.insert_link(link_label, link_url)

    def _append_status_link(self, label, url):
        self.status_text.insert("end", "\n")
        self.status_text.insert_link(label, url)

    def _resolve_options(self):
        exact = self.exact_match_var.get()
        all_terms = self.all_terms_var.get()
        mode = ArtpriceURLBuilder.resolve_mode(exact, all_terms)
        self.log(f"Search mode resolved to: {mode}")
        return exact, all_terms

    def choose_image_file(self):
        path = filedialog.askopenfilename(
            title="Select Auction Listing Screenshot",
            filetypes=[("Image Files", "*.png *.jpg *.jpeg *.tif *.tiff *.bmp *.webp")],
        )
        if not path:
            return
        self.image_input_path = path
        self.image_file_var.set(path)
        self.log(f'Listing image chosen: "{os.path.basename(path)}"')
        self._set_status_message(f"Listing image selected: {path}")

    def generate_link_from_image(self):
        if self.lookup.df is None:
            messagebox.showwarning("Missing Artist IDs", "Default artist ID file could not be loaded.")
            return

        image_path = self.image_file_var.get().strip()
        if not image_path:
            messagebox.showwarning("Missing Image", "Please choose a listing image.")
            return

        self.log(f"Image Entered: {image_path}")
        exact, all_terms = self._resolve_options()

        try:
            listing = self.extractor.extract_from_image(image_path)
            artist_id = self.lookup.lookup(listing.artist)

            self.artist_var.set(listing.artist)
            self.title_var.set(listing.title)

            keyword = ArtpriceURLBuilder.build_keyword(listing.title, exact, all_terms)

            if not artist_id:
                link = ArtpriceURLBuilder.build_url_without_artist(listing.title, exact, all_terms)
                self.artist_id_var.set("")
                self.keyword_var.set(keyword)
                self.last_generated_link = link
                self.output.delete("1.0", "end")
                self.output.insert("end", link)
                self.log(f"Missing artist ID match in single image mode: {listing.artist}")
                self.log(f"Generated fallback Artprice link without artist ID: {link}")
                self._set_status_message(
                    f"No artist ID found for: {listing.artist}. Generated fallback link without artist filter.",
                    "Open generated link in browser",
                    link,
                )
                return

            link = ArtpriceURLBuilder.build_url(artist_id, listing.title, exact, all_terms)
            self.artist_id_var.set(str(artist_id))
            self.keyword_var.set(keyword)
            self.last_generated_link = link
            self.output.delete("1.0", "end")
            self.output.insert("end", link)
            self.log(f"Single image extracted artist: {listing.artist}")
            self.log(f"Single image extracted title: {listing.title}")
            self.log(f"Generated Artprice link: {link}")
            self._set_status_message("Artprice link generated successfully.", "Open generated link in browser", link)

        except Exception as exc:
            self.log(f"Single image generation error: {exc}")
            messagebox.showerror("Error", str(exc))

    def choose_batch_file(self):
        path = filedialog.askopenfilename(title="Select Auction Spreadsheet", filetypes=[("Excel Workbook", "*.xlsx")])
        if not path:
            return
        self.batch_input_path = path
        self.batch_file_var.set(path)
        self.generate_batch_button.pack(anchor="w", pady=(0, 10))
        self.log(f'Auction Spreadsheet "{os.path.basename(path)}" chosen for link generations')
        self._set_status_message(f"Batch spreadsheet selected: {path}")

    def process_batch_file(self):
        if self.lookup.df is None:
            messagebox.showwarning("Missing Artist IDs", "Default artist ID file could not be loaded.")
            return

        input_path = self.batch_file_var.get().strip()
        if not input_path:
            messagebox.showwarning("Missing Spreadsheet", "Choose an XLSX file to process.")
            return

        exact, all_terms = self._resolve_options()
        self.batch_log.delete("1.0", "end")
        self.log(f"Starting batch processing for: {input_path}")

        def progress(row_num, total, status):
            line = f"Row {row_num}/{total}: {status}"
            self.batch_log.insert("end", line + "\n")
            self.batch_log.see("end")
            self.master.update_idletasks()

        try:
            output_path, missing_path, artist_col, title_col, result_df = self.batch_processor.process_workbook(
                input_path=input_path,
                exact_match=exact,
                all_terms=all_terms,
                progress_callback=progress,
            )
            self.last_exported_file = output_path
            self.last_missing_file = missing_path
            ok_count = int((result_df[STATUS_COLUMN] == "OK").sum())
            missing_count = int((result_df[STATUS_COLUMN] == UNKNOWN_ARTIST_MESSAGE).sum())
            self.log(f"Batch done. OK rows: {ok_count}. Missing artist IDs: {missing_count}.")
            if missing_path:
                self.log(f"Missing Artist Artprice IDs.xlsx exported to desktop: {missing_path}")
            self._set_status_message(
                f"Updated spreadsheet exported. Artist column: {artist_col or 'None'} | Title column: {title_col or 'None'} | OK: {ok_count} | Missing: {missing_count}",
                "Open exported spreadsheet",
                f"file://{os.path.abspath(output_path)}",
            )
            if missing_path:
                self._append_status_link("Open Missing Artist Artprice IDs.xlsx", f"file://{os.path.abspath(missing_path)}")
        except Exception as exc:
            self.log(f"Batch processing error: {exc}")
            messagebox.showerror("Batch Processing Error", str(exc))

    def open_link(self):
        if self.last_generated_link.startswith("http"):
            webbrowser.open(self.last_generated_link)
        else:
            messagebox.showinfo("Nothing to Open", "Generate a valid Artprice link first.")

    def copy_link(self):
        text = self.last_generated_link.strip()
        if not text:
            messagebox.showinfo("Nothing to Copy", "Generate a link first.")
            return
        self.master.clipboard_clear()
        self.master.clipboard_append(text)
        self.master.update()
        self.log("Generated Artprice link copied to clipboard")
        self._set_status_message("Generated Artprice link copied to clipboard.")

    def clear_image_tab(self):
        self.image_file_var.set("")
        self.artist_var.set("")
        self.title_var.set("")
        self.artist_id_var.set("")
        self.keyword_var.set("")
        self.last_generated_link = ""
        self.output.delete("1.0", "end")
        self.log("Single image panel cleared")
        self._set_status_message("Single image panel cleared.")

    def _animate_banner(self, tick):
        w = max(self.banner.winfo_width(), 900)
        h = int(self.banner["height"])
        self.banner.delete("all")
        steps = 78
        pink_pass = ((tick // 28) % 3) == 2

        for i in range(steps):
            x0 = i * w / steps
            x1 = (i + 1) * w / steps + 1
            phase = (i / steps * 2 * math.pi) + tick / 8
            glow = (math.sin(phase) + 1) / 2
            base = blend_hex(GOLD_1, GOLD_4, glow)
            color = base
            if pink_pass:
                pink_mix = max(0.0, math.sin(phase + 0.8))
                if pink_mix > 0:
                    color = blend_hex(base, blend_hex(PINK_1, PINK_2, glow), min(0.28, pink_mix * 0.28))
            self.banner.create_rectangle(x0, 0, x1, h, outline="", fill=color)

        self.banner.create_line(0, h - 1, w, h - 1, fill="#d0b06a", width=1)
        self.master.after(85, lambda: self._animate_banner(tick + 1))


def format_output_workbook(path: str, link_header: str):
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="D9C07D")
    zebra_fill = PatternFill("solid", fgColor="FAF6EE")
    ok_fill = PatternFill("solid", fgColor="DCFCE7")
    warn_fill = PatternFill("solid", fgColor="FCE7F3")
    thin = Side(style="thin", color="D6C29A")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="2A2118")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    link_col = None
    status_col = None
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(1, c).value or "")
        if header == link_header:
            link_col = c
        if header == STATUS_COLUMN:
            status_col = c

    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = zebra_fill

        if link_col:
            cell = ws.cell(row, link_col)
            if cell.value and str(cell.value).startswith("http"):
                url = str(cell.value)
                cell.hyperlink = url
                cell.value = "Open Link"
                cell.style = "Hyperlink"

        if status_col:
            status_cell = ws.cell(row, status_col)
            if str(status_cell.value) == "OK":
                status_cell.fill = ok_fill
            elif str(status_cell.value) == UNKNOWN_ARTIST_MESSAGE:
                status_cell.fill = warn_fill

    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            length = max(length, len(str(value)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(length + 3, 14), 48)

    wb.save(path)


def main():
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(0)