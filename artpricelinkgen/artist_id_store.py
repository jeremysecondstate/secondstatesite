from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


class ArtistIdWorkbookStore:
    """Append resolved Artist/ID rows to the configured Artist IDs workbook.

    A timestamped backup is created before the first write in each process.
    """

    def __init__(self, workbook_path: str, logger=None):
        self.workbook_path = Path(workbook_path)
        self.logger = logger or (lambda message: None)
        self._backup_path: Path | None = None

    @property
    def can_write(self) -> bool:
        return self.workbook_path.exists() and self.workbook_path.suffix.lower() == ".xlsx"

    def backup_once(self) -> Path:
        if self._backup_path is not None:
            return self._backup_path

        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Artist ID workbook does not exist: {self.workbook_path}")
        if self.workbook_path.suffix.lower() != ".xlsx":
            raise ValueError("Automatic Artist ID writes require an .xlsx workbook.")

        stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        backup_name = f"{self.workbook_path.stem}.backup_{stamp}{self.workbook_path.suffix}"
        backup_path = self.workbook_path.with_name(backup_name)
        shutil.copy2(self.workbook_path, backup_path)
        self._backup_path = backup_path
        self.logger(f"Backed up Artist IDs workbook: {backup_path}")
        return backup_path

    def append_artist_id(self, artist: str, artist_id: str, source_url: str = "", confidence: str = "") -> bool:
        artist = str(artist or "").strip()
        artist_id = str(artist_id or "").strip()
        if not artist or not artist_id:
            return False
        if not self.can_write:
            self.logger(f"Cannot auto-save Artist ID; workbook is missing or is not .xlsx: {self.workbook_path}")
            return False

        self.backup_once()

        wb = load_workbook(self.workbook_path)
        ws = wb.active

        headers = [str(cell.value or "").strip() for cell in ws[1]]
        header_map = {header.lower(): idx + 1 for idx, header in enumerate(headers) if header}

        artist_col = header_map.get("artist")
        id_col = header_map.get("id")
        if not artist_col or not id_col:
            raise ValueError("Artist IDs workbook must contain 'Artist' and 'ID' columns.")

        # Avoid duplicate rows if the ID was already added earlier.
        normalized_artist = artist.lower().strip()
        for row in range(2, ws.max_row + 1):
            existing_artist = str(ws.cell(row, artist_col).value or "").lower().strip()
            existing_id = str(ws.cell(row, id_col).value or "").strip()
            if existing_artist == normalized_artist or existing_id == artist_id:
                self.logger(f"Artist ID already present, skipping append: {artist} / {artist_id}")
                return False

        next_row = ws.max_row + 1
        ws.cell(next_row, artist_col).value = artist
        ws.cell(next_row, id_col).value = artist_id

        # Optional metadata columns are filled only if they already exist, so the
        # main workbook can stay as a simple Artist/ID table.
        optional_values = {
            "resolved source url": source_url,
            "resolver confidence": confidence,
            "resolved at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        for header, value in optional_values.items():
            col = header_map.get(header)
            if col and value:
                ws.cell(next_row, col).value = value

        wb.save(self.workbook_path)
        self.logger(f"Added resolved Artist ID to workbook: {artist} / {artist_id}")
        return True
