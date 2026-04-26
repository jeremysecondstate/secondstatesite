from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from artpricelinkgen_v2.config import STATUS_COLUMN, UNKNOWN_ARTIST_MESSAGE


def format_output_workbook(path: str, link_header: str):
    wb = load_workbook(path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="D9C07D")
    zebra_fill = PatternFill("solid", fgColor="FAF6EE")
    ok_fill = PatternFill("solid", fgColor="DCFCE7")
    warn_fill = PatternFill("solid", fgColor="FCE7F3")
    thin = Side(style="thin", color="D6C29A")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="2A2118")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    link_col = None
    status_col = None
    for c in range(1, ws.max_column + 1):
        header = str(ws.cell(1, c).value or "")
        if header == link_header:
            link_col = c
        if header == STATUS_COLUMN:
            status_col = c

    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).fill = zebra_fill

        if link_col:
            cell = ws.cell(row, link_col)
            if cell.value and str(cell.value).startswith("http"):
                url = str(cell.value)
                cell.hyperlink = url
                cell.value = "Open Link"
                cell.style = "Hyperlink"

        if status_col:
            status_cell = ws.cell(row, status_col)
            if str(status_cell.value) == "OK":
                status_cell.fill = ok_fill
            elif str(status_cell.value) == UNKNOWN_ARTIST_MESSAGE:
                status_cell.fill = warn_fill

    for col_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            value = cell.value
            if value is None:
                continue
            length = max(length, len(str(value)))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(length + 3, 14), 48)

    wb.save(path)
