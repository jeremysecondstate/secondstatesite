import datetime as dt
import os
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


PARTNERS = ("Alex", "Jeremy", "Oliver")
PARTNER_LOOKUP = {name.lower(): name for name in PARTNERS}
WORKBOOK_FILENAME = "SUPREME.xlsx"


def _base_dir():
    return Path(__file__).resolve().parent.parent


def _private_data_root():
    configured = os.environ.get("SUPREME_PRIVATE_DATA_ROOT")
    if configured:
        return Path(configured)
    render_disk_root = Path("/var/data")
    if render_disk_root.exists():
        return render_disk_root / "private"
    return _base_dir() / "private_data"


def default_saved_workbook_path():
    configured = os.environ.get("SUPREME_WORKBOOK_PATH")
    if configured:
        return Path(configured)
    return _private_data_root() / "supreme" / WORKBOOK_FILENAME


def saved_workbook_info(path=None):
    workbook_path = Path(path) if path else default_saved_workbook_path()
    exists = workbook_path.exists()
    modified_at = None
    if exists:
        modified_at = dt.datetime.fromtimestamp(workbook_path.stat().st_mtime).strftime("%Y-%m-%d %I:%M %p")
    return {
        "path": str(workbook_path),
        "name": workbook_path.name,
        "exists": exists,
        "modified_at": modified_at,
    }


def save_uploaded_workbook(uploaded_file, path=None):
    filename = getattr(uploaded_file, "name", "") or ""
    if filename and not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("Please upload an Excel workbook ending in .xlsx or .xlsm.")

    workbook_path = Path(path) if path else default_saved_workbook_path()
    workbook_path.parent.mkdir(parents=True, exist_ok=True)

    with workbook_path.open("wb") as destination:
        if hasattr(uploaded_file, "chunks"):
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        else:
            destination.write(uploaded_file.read())
    return workbook_path


def _clean(value):
    return str(value or "").strip()


def _norm(value):
    return _clean(value).lower().replace("\n", " ").replace("  ", " ")


def _cell(row, index, default=None):
    if index is None:
        return default
    try:
        if index < 0:
            return default
        return row[index]
    except (IndexError, TypeError):
        return default


def _money(value):
    if value is None or value == "":
        return Decimal("0")
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = _clean(value).replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
    if not text or text.lower() in {"amount", "amount ($)", "low", "high", "total"}:
        return Decimal("0")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return Decimal("0")


def _money_float(value):
    return float(value.quantize(Decimal("0.01")))


def _money_display(value):
    return f"${_money_float(value):,.2f}"


def _date_bucket(value):
    if isinstance(value, dt.datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, dt.date):
        return value.strftime("%Y-%m")
    text = _clean(value)
    if not text:
        return "Unknown"
    try:
        return dt.datetime.fromisoformat(text).strftime("%Y-%m")
    except ValueError:
        return text[:7] if len(text) >= 7 else text


def _summary_period(title, fallback_index):
    text = _clean(title).upper()
    for period in ("JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE", "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER"):
        if period in text:
            return period.title()
    return f"Summary {fallback_index}"


def _find_contribution_header(ws):
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [_norm(cell) for cell in row]
        if "partner" in normalized and "type" in normalized and any("amount" in cell for cell in normalized):
            return row_index, normalized
    return None, []


def _column_map(headers):
    mapping = {}
    for idx, header in enumerate(headers):
        if header == "date":
            mapping["date"] = idx
        elif header == "partner":
            mapping["partner"] = idx
        elif header == "type":
            mapping["type"] = idx
        elif "amount" in header:
            mapping["amount"] = idx
        elif header == "notes":
            mapping["notes"] = idx
    return mapping


def _parse_contributions(ws):
    header_row, headers = _find_contribution_header(ws)
    if not header_row:
        return []
    columns = _column_map(headers)
    if "amount" not in columns:
        return []

    date_col = columns.get("date")
    partner_col = columns.get("partner")
    type_col = columns.get("type")
    amount_col = columns.get("amount")
    notes_col = columns.get("notes")

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        partner_raw = _clean(_cell(row, partner_col, ""))
        type_raw = _clean(_cell(row, type_col, ""))
        if _norm(partner_raw) == "partner" or _norm(type_raw) == "type":
            continue
        partner = PARTNER_LOOKUP.get(partner_raw.lower())
        if not partner:
            continue
        amount = _money(_cell(row, amount_col))
        if amount == 0 and not type_raw:
            continue
        date_value = _cell(row, date_col)
        rows.append(
            {
                "date": date_value,
                "month": _date_bucket(date_value),
                "partner": partner,
                "type": type_raw or "Uncategorized",
                "amount": amount,
                "notes": _clean(_cell(row, notes_col, "")),
            }
        )
    return rows


def _summary_title_above(ws, header_row_idx, start_col_idx):
    min_row = max(1, header_row_idx - 4)
    for row_idx in range(header_row_idx - 1, min_row - 1, -1):
        values = next(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True), ())
        for col_idx in range(max(0, start_col_idx - 1), min(len(values), start_col_idx + 4)):
            text = _clean(_cell(values, col_idx, ""))
            if "liquidation payout summary" in text.lower():
                return text
    return "Liquidation Payout Summary"


def _parse_liquidation_summaries(ws):
    summaries = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [_norm(cell) for cell in row]
        for col_idx, cell in enumerate(normalized):
            if cell == "partner" and any("deal capital" in item for item in normalized[col_idx : col_idx + 5]):
                header_map = {}
                for offset, header in enumerate(normalized[col_idx : col_idx + 5]):
                    absolute = col_idx + offset
                    if header == "partner":
                        header_map["partner"] = absolute
                    elif "deal capital" in header:
                        header_map["deal_capital"] = absolute
                    elif "profit" in header or "retained" in header:
                        header_map["retained_earnings"] = absolute
                    elif "liquidation" in header or "payout" in header:
                        header_map["liquidation_payout"] = absolute

                title = _summary_title_above(ws, row_idx, col_idx)
                summary = {"_title": title, "_period": _summary_period(title, len(summaries) + 1)}
                has_explicit_payout = header_map.get("liquidation_payout") is not None
                retained_total = Decimal("0")
                for data_row in ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 12, values_only=True):
                    partner = PARTNER_LOOKUP.get(_clean(_cell(data_row, header_map.get("partner"), "")).lower())
                    if not partner:
                        continue
                    deal_capital = _money(_cell(data_row, header_map.get("deal_capital")))
                    retained = _money(_cell(data_row, header_map.get("retained_earnings")))
                    payout = _money(_cell(data_row, header_map.get("liquidation_payout"))) if has_explicit_payout else deal_capital + retained
                    retained_total += retained
                    summary[partner] = {
                        "deal_capital": deal_capital,
                        "retained_earnings": retained,
                        "liquidation_payout": payout,
                    }
                if any(partner in summary for partner in PARTNERS):
                    summary["_retained_total"] = retained_total
                    summaries.append(summary)
    return summaries


def _build_profit_breakdown(summaries):
    period_summaries = summaries[-2:]
    period_labels = [summary.get("_period", f"Summary {idx + 1}") for idx, summary in enumerate(period_summaries)]

    partner_rows = []
    for partner in PARTNERS:
        profits = []
        total = Decimal("0")
        for summary in period_summaries:
            amount = summary.get(partner, {}).get("retained_earnings", Decimal("0"))
            total += amount
            profits.append({"label": summary.get("_period", "Profit"), "amount": _money_display(amount), "value": _money_float(amount)})
        partner_rows.append({"partner": partner, "profits": profits, "total": _money_display(total), "total_value": _money_float(total)})

    company_profit_rows = []
    company_total = Decimal("0")
    for summary in period_summaries:
        amount = summary.get("_retained_total", Decimal("0"))
        company_total += amount
        company_profit_rows.append({"label": summary.get("_period", "Profit"), "amount": _money_display(amount), "value": _money_float(amount)})

    return {
        "period_labels": period_labels,
        "partner_rows": partner_rows,
        "company_profit_rows": company_profit_rows,
        "company_total": _money_display(company_total),
        "company_total_value": _money_float(company_total),
    }


def build_dashboard(workbook_file=None, workbook_path=None):
    source_name = "SUPREME workbook"
    source_info = None
    if workbook_file:
        source_name = getattr(workbook_file, "name", source_name)
        wb = load_workbook(workbook_file, data_only=True, read_only=True)
    else:
        workbook_path = Path(workbook_path) if workbook_path else default_saved_workbook_path()
        if not workbook_path.exists():
            raise FileNotFoundError("Upload and save SUPREME.xlsx, or set SUPREME_WORKBOOK_PATH on the server.")
        source_name = workbook_path.name
        source_info = saved_workbook_info(workbook_path)
        wb = load_workbook(workbook_path, data_only=True, read_only=True)

    sheet_name = "Contributions" if "Contributions" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = _parse_contributions(ws)
    liquidation_summaries = _parse_liquidation_summaries(ws)
    liquidation_summary = liquidation_summaries[-1] if liquidation_summaries else {}
    profit_breakdown = _build_profit_breakdown(liquidation_summaries)

    totals_by_partner = defaultdict(Decimal)
    deal_capital_by_partner = defaultdict(Decimal)
    expenses_by_partner = defaultdict(Decimal)
    totals_by_type = defaultdict(Decimal)
    monthly_totals = defaultdict(Decimal)

    for row in rows:
        amount = row["amount"]
        partner = row["partner"]
        type_label = row["type"] or "Uncategorized"
        type_lower = type_label.lower()
        totals_by_partner[partner] += amount
        totals_by_type[type_label] += amount
        monthly_totals[row["month"]] += amount
        if "deal capital" in type_lower:
            deal_capital_by_partner[partner] += amount
        elif "expense" in type_lower or "infrastructure" in type_lower:
            expenses_by_partner[partner] += amount

    partner_cards = []
    for partner in PARTNERS:
        summary = liquidation_summary.get(partner, {})
        deal_capital = summary.get("deal_capital", deal_capital_by_partner[partner])
        retained = summary.get("retained_earnings", Decimal("0"))
        payout = summary.get("liquidation_payout", deal_capital + retained)
        partner_cards.append(
            {
                "partner": partner,
                "deal_capital": _money_display(deal_capital),
                "deal_capital_value": _money_float(deal_capital),
                "retained_earnings": _money_display(retained),
                "retained_earnings_value": _money_float(retained),
                "liquidation_payout": _money_display(payout),
                "liquidation_payout_value": _money_float(payout),
                "total_contributions": _money_display(totals_by_partner[partner]),
                "expense_total": _money_display(expenses_by_partner[partner]),
            }
        )

    type_rows = [
        {"type": label, "amount": _money_display(amount), "value": _money_float(amount)}
        for label, amount in sorted(totals_by_type.items(), key=lambda item: abs(item[1]), reverse=True)
    ]
    month_rows = [
        {"month": month, "amount": _money_display(amount), "value": _money_float(amount)}
        for month, amount in sorted(monthly_totals.items())
    ]

    chart_data = {
        "partners": [card["partner"] for card in partner_cards],
        "dealCapital": [card["deal_capital_value"] for card in partner_cards],
        "retainedEarnings": [card["retained_earnings_value"] for card in partner_cards],
        "liquidationPayout": [card["liquidation_payout_value"] for card in partner_cards],
        "types": [row["type"] for row in type_rows],
        "typeAmounts": [row["value"] for row in type_rows],
        "months": [row["month"] for row in month_rows],
        "monthAmounts": [row["value"] for row in month_rows],
        "profitPeriods": profit_breakdown["period_labels"],
        "companyProfits": [row["value"] for row in profit_breakdown["company_profit_rows"]],
    }

    total_deal_capital = sum(Decimal(str(card["deal_capital_value"])) for card in partner_cards)
    total_retained = sum(Decimal(str(card["retained_earnings_value"])) for card in partner_cards)
    total_payout = sum(Decimal(str(card["liquidation_payout_value"])) for card in partner_cards)

    return {
        "source_name": source_name,
        "source_info": source_info,
        "summary_title": liquidation_summary.get("_title", "Liquidation Payout Summary"),
        "sheet_name": sheet_name,
        "row_count": len(rows),
        "partner_cards": partner_cards,
        "profit_breakdown": profit_breakdown,
        "type_rows": type_rows[:12],
        "month_rows": month_rows[-18:],
        "totals": {
            "deal_capital": _money_display(total_deal_capital),
            "retained_earnings": _money_display(total_retained),
            "liquidation_payout": _money_display(total_payout),
            "transactions": len(rows),
        },
        "chart_data": chart_data,
    }
