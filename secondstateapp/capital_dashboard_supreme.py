import re
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from .capital_dashboard import (
    PARTNERS,
    PARTNER_LOOKUP,
    _cell,
    _clean,
    _column_map,
    _date_bucket,
    _money,
    _money_display,
    _money_float,
    _norm,
    _parse_contributions,
    _parse_liquidation_summaries,
    default_saved_workbook_path,
    saved_workbook_info,
)


MONTHS = (
    "JANUARY",
    "FEBRUARY",
    "MARCH",
    "APRIL",
    "MAY",
    "JUNE",
    "JULY",
    "AUGUST",
    "SEPTEMBER",
    "OCTOBER",
    "NOVEMBER",
    "DECEMBER",
)
MONTH_NUMBER = {month: index for index, month in enumerate(MONTHS, start=1)}


def _month_in_text(text):
    upper = _clean(text).upper()
    matches = [(upper.find(month), month) for month in MONTHS if upper.find(month) >= 0]
    return min(matches, default=(None, None))[1]


def _period_label(text, fallback_index):
    upper = _clean(text).upper()
    month = _month_in_text(upper)
    year_match = re.search(r"\b(20\d{2})\b", upper)
    if month and year_match:
        return f"{month.title()} {year_match.group(1)}"
    if month:
        return month.title()
    return f"Sale Period {fallback_index}"


def _period_sort_key(text, fallback_index):
    upper = _clean(text).upper()
    month = _month_in_text(upper)
    year_match = re.search(r"\b(20\d{2})\b", upper)
    year = int(year_match.group(1)) if year_match else 0
    return (year, MONTH_NUMBER.get(month, 0), fallback_index)


def _find_sales_header(ws):
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [_norm(cell) for cell in row]
        if (
            "artist" in normalized
            and "profit" in normalized
            and any("net sale price" in item for item in normalized)
            and any("sold hammer price" in item for item in normalized)
        ):
            return row_index, normalized
    return None, []


def _sales_column_map(headers):
    mapping = {}
    for idx, header in enumerate(headers):
        if header == "date":
            mapping["date"] = idx
        elif header == "artist":
            mapping["artist"] = idx
        elif header in {"name", "piece", "title"}:
            mapping["piece"] = idx
        elif "sale location" in header:
            mapping["sale_location"] = idx
        elif header == "sold hammer price $":
            mapping["gross_sales"] = idx
        elif header in {"commission $", "comission $"}:
            mapping["commission"] = idx
        elif header == "net sale price $":
            mapping["net_sales"] = idx
        elif header == "purchase price $":
            mapping["purchase_basis"] = idx
        elif header == "restoration cost $":
            mapping["restoration_cost"] = idx
        elif header == "profit":
            mapping["profit"] = idx
    return mapping


def _parse_sales_sheet(ws, fallback_index):
    header_row, headers = _find_sales_header(ws)
    if not header_row:
        return None
    columns = _sales_column_map(headers)
    if not {"artist", "piece", "net_sales", "profit"}.issubset(columns):
        return None

    banner_text = " ".join(
        _clean(cell)
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True)
        for cell in row
        if _clean(cell)
    )
    period_source = f"{ws.title} {banner_text}"
    period = _period_label(period_source, fallback_index)
    sales = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        artist = _clean(_cell(row, columns.get("artist"), ""))
        piece = _clean(_cell(row, columns.get("piece"), ""))
        if not artist or not piece:
            continue
        gross_sales = _money(_cell(row, columns.get("gross_sales")))
        net_sales = _money(_cell(row, columns.get("net_sales")))
        profit = _money(_cell(row, columns.get("profit")))
        if gross_sales == 0 and net_sales == 0 and profit == 0:
            continue
        sales.append(
            {
                "period": period,
                "artist": artist,
                "piece": piece,
                "sale_location": _clean(_cell(row, columns.get("sale_location"), "")),
                "gross_sales": gross_sales,
                "commission": _money(_cell(row, columns.get("commission"))),
                "net_sales": net_sales,
                "purchase_basis": _money(_cell(row, columns.get("purchase_basis"))),
                "restoration_cost": _money(_cell(row, columns.get("restoration_cost"))),
                "profit": profit,
            }
        )

    if not sales:
        return None
    return {
        "period": period,
        "sheet_name": ws.title,
        "sort_key": _period_sort_key(period_source, fallback_index),
        "sale_count": len(sales),
        "gross_sales_value": sum((sale["gross_sales"] for sale in sales), Decimal("0")),
        "commissions_value": sum((sale["commission"] for sale in sales), Decimal("0")),
        "net_sales_value": sum((sale["net_sales"] for sale in sales), Decimal("0")),
        "purchase_basis_value": sum((sale["purchase_basis"] for sale in sales), Decimal("0")),
        "restoration_costs_value": sum((sale["restoration_cost"] for sale in sales), Decimal("0")),
        "profit_value": sum((sale["profit"] for sale in sales), Decimal("0")),
        "sales": sales,
    }


def _parse_sales(wb):
    summaries = []
    for index, sheet_name in enumerate(wb.sheetnames, start=1):
        if "sales" not in _norm(sheet_name):
            continue
        summary = _parse_sales_sheet(wb[sheet_name], index)
        if summary:
            summaries.append(summary)
    return sorted(summaries, key=lambda row: row["sort_key"])


def _parse_logged_payouts(wb):
    sheet_name = next((name for name in ("Company Payouts v2", "Company Payouts") if name in wb.sheetnames), None)
    by_partner = defaultdict(Decimal)
    entries = []
    if not sheet_name:
        return {"sheet_name": None, "by_partner": by_partner, "entries": entries, "total": Decimal("0")}

    ws = wb[sheet_name]
    header_row = None
    headers = []
    for row_index, row in enumerate(ws.iter_rows(values_only=True), start=1):
        normalized = [_norm(cell) for cell in row]
        if "partner" in normalized and any("amount" in item for item in normalized) and "type" in normalized:
            header_row, headers = row_index, normalized
            break
    if not header_row:
        return {"sheet_name": sheet_name, "by_partner": by_partner, "entries": entries, "total": Decimal("0")}

    columns = _column_map(headers)
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        partner = PARTNER_LOOKUP.get(_clean(_cell(row, columns.get("partner"), "")).lower())
        if not partner:
            continue
        amount = _money(_cell(row, columns.get("amount")))
        if amount == 0:
            continue
        entries.append({"partner": partner, "amount": amount})
        by_partner[partner] += amount

    return {
        "sheet_name": sheet_name,
        "by_partner": by_partner,
        "entries": entries,
        "total": sum(by_partner.values(), Decimal("0")),
    }


def _build_profit_breakdown(summaries):
    labels = [summary.get("_period", f"Summary {index + 1}") for index, summary in enumerate(summaries)]
    partner_rows = []
    partner_totals = {}
    for partner in PARTNERS:
        profits = []
        total = Decimal("0")
        for summary in summaries:
            amount = summary.get(partner, {}).get("retained_earnings", Decimal("0"))
            total += amount
            profits.append({"label": summary.get("_period", "Profit"), "amount": _money_display(amount), "value": _money_float(amount)})
        partner_totals[partner] = total
        partner_rows.append({"partner": partner, "profits": profits, "total": _money_display(total), "total_value": _money_float(total)})

    company_rows = []
    company_total = Decimal("0")
    for summary in summaries:
        amount = summary.get("_retained_total", Decimal("0"))
        company_total += amount
        company_rows.append({"label": summary.get("_period", "Profit"), "amount": _money_display(amount), "value": _money_float(amount)})

    return {
        "title": " + ".join(labels) + " Profit Detail" if labels else "Profit Detail",
        "period_labels": labels,
        "partner_rows": partner_rows,
        "partner_totals": partner_totals,
        "company_profit_rows": company_rows,
        "company_total": _money_display(company_total),
        "company_total_value": _money_float(company_total),
        "company_total_decimal": company_total,
    }


def _build_sales_dashboard(summaries):
    totals = {
        key: sum((row[key] for row in summaries), Decimal("0"))
        for key in (
            "gross_sales_value",
            "commissions_value",
            "net_sales_value",
            "purchase_basis_value",
            "restoration_costs_value",
            "profit_value",
        )
    }
    periods = []
    all_sales = []
    for row in summaries:
        periods.append(
            {
                "period": row["period"],
                "sheet_name": row["sheet_name"],
                "sale_count": row["sale_count"],
                "gross_sales": _money_display(row["gross_sales_value"]),
                "net_sales": _money_display(row["net_sales_value"]),
                "net_sales_value": _money_float(row["net_sales_value"]),
                "profit": _money_display(row["profit_value"]),
                "profit_value": _money_float(row["profit_value"]),
            }
        )
        all_sales.extend(row["sales"])

    recent_sales = [
        {
            "period": sale["period"],
            "artist": sale["artist"],
            "piece": sale["piece"],
            "sale_location": sale["sale_location"],
            "net_sales": _money_display(sale["net_sales"]),
            "profit": _money_display(sale["profit"]),
        }
        for sale in reversed(all_sales)
    ][:8]

    return {
        "periods": periods,
        "recent_sales": recent_sales,
        "sheet_count": len(periods),
        "sale_count": sum(row["sale_count"] for row in summaries),
        "gross_sales": _money_display(totals["gross_sales_value"]),
        "net_sales": _money_display(totals["net_sales_value"]),
        "profit": _money_display(totals["profit_value"]),
        "profit_decimal": totals["profit_value"],
    }


def build_dashboard(workbook_file=None, workbook_path=None):
    source_name = "SUPREME workbook"
    source_info = None
    if workbook_file:
        source_name = getattr(workbook_file, "name", source_name)
        wb = load_workbook(workbook_file, data_only=True, read_only=True)
    else:
        workbook_path = Path(workbook_path) if workbook_path else default_saved_workbook_path()
        if not workbook_path.exists():
            raise FileNotFoundError("Upload and save SUPREME.xlsx, or set SUPREME_WORKBOOK_PATH on the server.")
        source_name = workbook_path.name
        source_info = saved_workbook_info(workbook_path)
        wb = load_workbook(workbook_path, data_only=True, read_only=True)

    sheet_name = "Contributions" if "Contributions" in wb.sheetnames else wb.sheetnames[0]
    rows = _parse_contributions(wb[sheet_name])
    summaries = _parse_liquidation_summaries(wb[sheet_name])
    latest_summary = summaries[-1] if summaries else {}
    profit_breakdown = _build_profit_breakdown(summaries)
    sales = _build_sales_dashboard(_parse_sales(wb))
    logged_payouts = _parse_logged_payouts(wb)

    totals_by_partner = defaultdict(Decimal)
    deal_capital_by_partner = defaultdict(Decimal)
    expenses_by_partner = defaultdict(Decimal)
    totals_by_type = defaultdict(Decimal)
    monthly_totals = defaultdict(Decimal)
    for row in rows:
        amount = row["amount"]
        partner = row["partner"]
        type_label = row["type"] or "Uncategorized"
        type_lower = type_label.lower()
        totals_by_partner[partner] += amount
        totals_by_type[type_label] += amount
        monthly_totals[row["month"]] += amount
        if "deal capital" in type_lower:
            deal_capital_by_partner[partner] += amount
        elif any(label in type_lower for label in ("expense", "infrastructure", "maintenance")):
            expenses_by_partner[partner] += amount

    partner_cards = []
    total_deal_capital = Decimal("0")
    latest_period_profit = Decimal("0")
    total_payout = Decimal("0")
    for partner in PARTNERS:
        summary = latest_summary.get(partner, {})
        deal_capital = summary.get("deal_capital", deal_capital_by_partner[partner])
        retained = summary.get("retained_earnings", Decimal("0"))
        payout = summary.get("liquidation_payout", deal_capital + retained)
        cumulative_profit = profit_breakdown["partner_totals"].get(partner, Decimal("0"))
        logged_payout = logged_payouts["by_partner"].get(partner, Decimal("0"))
        total_deal_capital += deal_capital
        latest_period_profit += retained
        total_payout += payout
        partner_cards.append(
            {
                "partner": partner,
                "deal_capital": _money_display(deal_capital),
                "deal_capital_value": _money_float(deal_capital),
                "retained_earnings": _money_display(retained),
                "cumulative_profit": _money_display(cumulative_profit),
                "liquidation_payout": _money_display(payout),
                "liquidation_payout_value": _money_float(payout),
                "logged_payouts": _money_display(logged_payout),
                "total_contributions": _money_display(totals_by_partner[partner]),
                "expense_total": _money_display(expenses_by_partner[partner]),
            }
        )

    type_rows = [
        {"type": label, "amount": _money_display(amount), "value": _money_float(amount)}
        for label, amount in sorted(totals_by_type.items(), key=lambda item: abs(item[1]), reverse=True)
    ]
    month_rows = [
        {"month": month, "amount": _money_display(amount), "value": _money_float(amount)}
        for month, amount in sorted(monthly_totals.items())
    ]
    company_profit = sales["profit_decimal"] if sales["periods"] else profit_breakdown["company_total_decimal"]
    source_tabs = [sheet_name] + [period["sheet_name"] for period in sales["periods"]]

    return {
        "source_name": source_name,
        "source_info": source_info,
        "summary_title": latest_summary.get("_title", "Liquidation Payout Summary"),
        "latest_period": latest_summary.get("_period", "Latest"),
        "sheet_name": sheet_name,
        "source_tabs": source_tabs,
        "source_tab_count": len(source_tabs),
        "row_count": len(rows),
        "partner_cards": partner_cards,
        "profit_breakdown": profit_breakdown,
        "sales": sales,
        "logged_payouts": {
            "sheet_name": logged_payouts["sheet_name"],
            "entry_count": len(logged_payouts["entries"]),
            "total": _money_display(logged_payouts["total"]),
        },
        "type_rows": type_rows[:12],
        "month_rows": month_rows[-18:],
        "totals": {
            "deal_capital": _money_display(total_deal_capital),
            "latest_period_profit": _money_display(latest_period_profit),
            "company_profit": _money_display(company_profit),
            "net_sales": sales["net_sales"],
            "gross_sales": sales["gross_sales"],
            "liquidation_payout": _money_display(total_payout),
            "logged_payouts": _money_display(logged_payouts["total"]),
            "transactions": len(rows),
            "sales_count": sales["sale_count"],
        },
        "chart_data": {
            "partners": [card["partner"] for card in partner_cards],
            "dealCapital": [card["deal_capital_value"] for card in partner_cards],
            "liquidationPayout": [card["liquidation_payout_value"] for card in partner_cards],
            "types": [row["type"] for row in type_rows],
            "typeAmounts": [row["value"] for row in type_rows],
            "months": [row["month"] for row in month_rows],
            "monthAmounts": [row["value"] for row in month_rows],
            "salesPeriods": [row["period"] for row in sales["periods"]],
            "netSales": [row["net_sales_value"] for row in sales["periods"]],
            "salesProfits": [row["profit_value"] for row in sales["periods"]],
        },
    }
