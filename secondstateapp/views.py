import json
import os
# from django.conf import settings
from django.core.files.storage import default_storage
from django.http import JsonResponse
from django.shortcuts import render
from django.views.decorators.csrf import csrf_exempt
from .models import Artwork, ArtworkImage, Sale  # Import ArtworkImage model
import tempfile
import pandas as pd
from django.conf import settings
import openpyxl


def _sales_path():
    return getattr(settings, "SALES_XLSX_PATH", "/var/data/sales/SUPREME SALES.xlsx")

def _load_sales_df():
    """
    Your file has title rows; the real header row contains Artist + Name.
    We detect it like you already do in catalogapp.py.
    """
    path = _sales_path()
    preview = pd.read_excel(path, header=None, nrows=40)

    header_row = None
    for i in range(len(preview)):
        row_vals = preview.iloc[i].astype(str).str.strip().tolist()
        if "Artist" in row_vals and "Name" in row_vals and "Date" in row_vals:
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find header row (Artist/Name/Date) in SUPREME SALES.xlsx")

    df = pd.read_excel(path, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]

    # Drop junk rows
    if "Artist" in df.columns:
        df = df[df["Artist"].notna()]
    return df, header_row

def _rewrite_sales_workbook(df, header_row):
    """
    Write back while preserving the title rows above header_row:
    easiest safe approach: load the original file with openpyxl and only overwrite the table region.
    """

    path = _sales_path()
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # clear old table rows (everything below header row)
    start_row = header_row + 2  # 1-based: header row + 1, then data begins next row
    max_row = ws.max_row
    if max_row >= start_row:
        ws.delete_rows(start_row, max_row - start_row + 1)

    # write df rows in the same column order as existing header
    headers = [cell.value for cell in ws[header_row + 1]]
    headers = [str(h).strip() if h is not None else "" for h in headers]

    # Ensure df has all columns
    for h in headers:
        if h and h not in df.columns:
            df[h] = ""

    df_out = df[headers] if all(h in df.columns for h in headers if h) else df

    for r_idx, row in enumerate(df_out.itertuples(index=False), start=start_row):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=None if (pd.isna(val)) else val)

    # atomic-ish write
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    os.replace(tmp_path, path)



def healthz(request):
    return JsonResponse({"ok": True})


def home(request):
    return render(request, "home.html")

def about(request):
    return render(request, "about.html")

def contact(request):
    return render(request, "contact.html")

def gallery(request):
    # Show the same content as /artworks/
    artworks = Artwork.objects.all()
    # artworks = Artwork.objects.order_by("-id")
    return render(request, "artworks/artwork_list.html", {"artworks": artworks})

def pieces_sold(request):
    sales = Sale.objects.order_by("-date", "-id")
    return render(request, "pieces_sold.html", {"sales": sales})

@csrf_exempt
def upload_sales_sheet(request):
    """
    Upload an Excel sales sheet, parse it, and upsert into the Sale table.
    """
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method"}, status=405)
    if not _authorized(request):
        return JsonResponse({"error": "Unauthorized"}, status=401)

    up = request.FILES.get("file")
    if not up:
        return JsonResponse({"error": "Missing file field 'file'."}, status=400)

    try:
        # Save upload to a temp file so pandas can read it safely
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            for chunk in up.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        # Detect header row like your desktop app
        preview = pd.read_excel(tmp_path, header=None, nrows=40)
        header_row = None
        for i in range(len(preview)):
            row_vals = preview.iloc[i].astype(str).str.strip().tolist()
            if "Artist" in row_vals and "Name" in row_vals and "Date" in row_vals:
                header_row = i
                break
        if header_row is None:
            raise ValueError("Could not find header row with Artist/Name/Date.")

        df = pd.read_excel(tmp_path, header=header_row)
        df.columns = [str(c).strip() for c in df.columns]

        # Normalize Name -> title field
        if "Name" in df.columns and "Title" not in df.columns:
            df["Title"] = df["Name"]

        # Drop blank rows
        if "Artist" in df.columns:
            df = df[df["Artist"].notna()]

        # Clear + reload (simple + safe). If you want true upsert later, we can do it.
        Sale.objects.all().delete()

        def parse_money(x):
            if pd.isna(x) or x == "":
                return None
            s = str(x).replace("$", "").replace(",", "").strip()
            try:
                return float(s)
            except Exception:
                return None

        created = 0
        for _, r in df.iterrows():
            # parse date
            raw_date = r.get("Date", None)
            date_val = pd.to_datetime(raw_date, errors="coerce")
            date_val = None if pd.isna(date_val) else date_val.date()

            Sale.objects.create(
                date=date_val,
                artist=(str(r.get("Artist", "")).strip() or None),
                title=(str(r.get("Title", r.get("Name", ""))).strip() or None),
                sale_location=(str(r.get("Sale Location", "")).strip() or None),
                net_sale_price=parse_money(r.get("Net Sale Price $", None)),
                auction_house=(str(r.get("Auction House", "")).strip() or None),
            )
            created += 1

        return JsonResponse({"message": "Imported sales", "count": created}, status=201)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass



def sales_list(request):
    try:
        df, _header_row = _load_sales_df()

        # Create a stable row id for delete operations
        # Spreadsheet has an 'Unnamed: 0' column that works well as an id column.
        id_col = "Unnamed: 0" if "Unnamed: 0" in df.columns else None
        if id_col:
            df[id_col] = df[id_col].fillna("")
        else:
            df["row_id"] = ""

        show_cols = [c for c in [
            "Unnamed: 0", "Date", "Artist", "Name", "Sale Location", "Sold Hammer Price $", "Net Sale Price $", "Profit"
        ] if c in df.columns]

        rows = df[show_cols].fillna("").to_dict(orient="records")
        return JsonResponse({"sales": rows})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

@csrf_exempt
def sales_add(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method"}, status=405)
    if not _authorized(request):
        return JsonResponse({"error": "Unauthorized"}, status=401)

    try:
        data = json.loads(request.body)

        df, header_row = _load_sales_df()

        # Assign next numeric id in Unnamed: 0 if present
        if "Unnamed: 0" in df.columns:
            used = pd.to_numeric(df["Unnamed: 0"], errors="coerce").dropna()
            next_id = int(used.max()) + 1 if len(used) else 1
            row_id = next_id
        else:
            row_id = None

        new_row = {}
        # Minimal required fields (match your spreadsheet headers)
        for col in df.columns:
            new_row[col] = ""

        if row_id is not None:
            new_row["Unnamed: 0"] = row_id

        # Map incoming JSON -> spreadsheet columns
        # (adjust these keys to match what you want to input in the desktop app)
        new_row["Date"] = data.get("date", "")
        new_row["Artist"] = data.get("artist", "")
        new_row["Name"] = data.get("name", "")
        if "Sale Location" in df.columns:
            new_row["Sale Location"] = data.get("sale_location", "")
        if "Sold Hammer Price $" in df.columns:
            new_row["Sold Hammer Price $"] = data.get("sold_hammer_price_usd", "")
        if "Net Sale Price $" in df.columns:
            new_row["Net Sale Price $"] = data.get("net_sale_price_usd", "")
        if "Profit" in df.columns:
            new_row["Profit"] = data.get("profit", "")

        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        _rewrite_sales_workbook(df, header_row)

        return JsonResponse({"message": "Sale added", "row_id": row_id}, status=201)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)

@csrf_exempt
def sales_delete(request):
    if request.method != "POST":
        return JsonResponse({"error": "Invalid request method"}, status=405)
    if not _authorized(request):
        return JsonResponse({"error": "Unauthorized"}, status=401)

    try:
        data = json.loads(request.body)
        row_id = str(data.get("row_id", "")).strip()
        if not row_id:
            return JsonResponse({"error": "row_id required"}, status=400)

        df, header_row = _load_sales_df()

        if "Unnamed: 0" not in df.columns:
            return JsonResponse({"error": "Workbook has no id column (Unnamed: 0)."}, status=400)

        before = len(df)
        df = df[df["Unnamed: 0"].astype(str).str.strip() != row_id]
        after = len(df)

        if before == after:
            return JsonResponse({"error": "Row not found"}, status=404)

        _rewrite_sales_workbook(df, header_row)
        return JsonResponse({"message": "Sale deleted"}, status=200)

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=400)


def _authorized(request):
    expected = os.environ.get("CATALOG_API_KEY")
    if not expected:
        return False
    return request.headers.get("X-API-KEY") == expected

def artwork_list(request):
    """Render artwork_list.html for normal browser requests; return JSON only when explicitly requested."""
    artworks = Artwork.objects.all()

    # Check if the request explicitly asks for JSON
    if 'format' in request.GET and request.GET['format'] == 'json':
        artwork_data = list(artworks.values('id', 'title', 'artist', 'description', 'price'))
        return JsonResponse({'artworks': artwork_data})

    # Otherwise, render the template
    return render(request, 'artworks/artwork_list.html', {'artworks': artworks})


def artwork_detail(request, pk):
    """Renders the detail page of an artwork from the database."""
    try:
        artwork = Artwork.objects.get(id=pk)  # Fetch from database
    except Artwork.DoesNotExist:
        return render(request, '404.html')  # Ensure you have a 404.html template

    return render(request, 'artworks/artwork_detail.html', {'artwork': artwork})


@csrf_exempt
def delete_artwork(request):
    """Deletes an artwork and all its associated images from the database and storage."""
    if request.method == 'POST':
        try:
            if not _authorized(request):
                return JsonResponse({"error": "Unauthorized"}, status=401)

            data = json.loads(request.body)
            title = data.get('title')

            # Find the artwork
            artwork = Artwork.objects.filter(title=title).first()
            if not artwork:
                return JsonResponse({'error': 'Artwork not found'}, status=404)

            # Delete all associated images from storage
            for image in artwork.images.all():
                # image_path = os.path.join(settings.MEDIA_ROOT, str(image.image))
                # if default_storage.exists(image_path):
                #     default_storage.delete(image_path)
                if image.image and default_storage.exists(image.image.name):
                    default_storage.delete(image.image.name)

            # Delete the artwork and its associated images from the database
            artwork.images.all().delete()  # Delete image entries
            artwork.delete()  # Delete artwork entry

            return JsonResponse({'message': f"'{title}' and all associated images deleted successfully!"}, status=200)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@csrf_exempt
def upload_artwork(request):
    if request.method == 'POST':
        try:
            if not _authorized(request):
                return JsonResponse({"error": "Unauthorized"}, status=401)

            data = request.POST

            # Save artwork to the database using raw fields from Excel
            artwork = Artwork.objects.create(
                title=data.get('title'),
                artist=data.get('artist'),
                year=data.get('year', ''),
                medium=data.get('medium', ''),
                paper_type=data.get('paper_type', ''),
                edition_size=data.get('edition_size', ''),
                printer=data.get('printer', ''),
                publisher=data.get('publisher', ''),
                dimensions_text=data.get('dimensions_text', ''),
                sheet_size=data.get('sheet_size', ''),
                catalog_number=data.get('catalog_number', ''),
                description=data.get('description', ''),  # ✅ This is now the 'Description/Notes' field
                price=float(data.get('price', 0)),
                is_available=True
            )

            # Save uploaded images
            for key, file in request.FILES.items():
                ArtworkImage.objects.create(artwork=artwork, image=file)

            return JsonResponse({'message': 'Artwork uploaded successfully!', 'id': artwork.id}, status=201)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Invalid request method'}, status=405)
